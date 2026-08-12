import io
from openpyxl import Workbook, load_workbook


def export_to_excel(headers, rows, filename='export.xlsx'):
    """
    导出数据到Excel
    headers: 列标题列表，如 ['名称', '分类', '数量']
    rows: 数据行列表，每行是列表或字典
    返回 (bytes_content, filename)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # 写表头
    ws.append(headers)

    # 写数据
    for row in rows:
        if isinstance(row, dict):
            # 如果是字典，按headers的key匹配（假设headers的key和字典key一致）
            # 这里简化处理：headers同时作为显示名和key
            ws.append([row.get(h, '') for h in headers])
        else:
            ws.append(list(row))

    # 调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue(), filename


def import_from_excel(file_storage):
    """
    从Excel文件导入数据
    file_storage: Flask的request.files中的文件对象
    返回 (headers_list, rows_list)，rows中每行是字典，key为表头
    """
    content = file_storage.read()
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []

    headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    data = []
    for row in rows[1:]:
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue  # 跳过空行
        row_dict = {}
        for i, header in enumerate(headers):
            if header:
                val = row[i] if i < len(row) else None
                row_dict[header] = val
        data.append(row_dict)

    return headers, data


def export_with_key_mapping(headers_mapping, rows, filename='export.xlsx'):
    """
    带字段映射的导出
    headers_mapping: 列表，每项为 (显示名, 字段key)，如 [('商品名称', 'name'), ('库存', 'current_stock')]
    rows: 字典列表
    """
    display_headers = [h[0] for h in headers_mapping]
    keys = [h[1] for h in headers_mapping]
    mapped_rows = [[row.get(k, '') for k in keys] for row in rows]
    return export_to_excel(display_headers, mapped_rows, filename)
